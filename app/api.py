import io
import json
from typing import Optional
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from . import calc, data_import, database as db, lingxing

router = APIRouter()


@router.get("/api/styles")
def list_styles():
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute("SELECT style_code, lifecycle FROM styles ORDER BY style_code")
    styles = [{"style_code": sc, "lifecycle": lc} for sc, lc in cur.fetchall()]
    cur.execute(
        "SELECT style_code, platform, account_name, account_prefix, stocking_days, quota FROM style_accounts ORDER BY style_code, sort_order"
    )
    accounts = [
        {
            "style_code": sc,
            "platform": p,
            "account_name": an,
            "account_prefix": ap,
            "stocking_days": sd,
            "quota": q,
        }
        for sc, p, an, ap, sd, q in cur.fetchall()
    ]
    conn.close()
    return {"styles": styles, "accounts": accounts}


@router.post("/api/styles/upload")
def upload_styles(file: UploadFile = File(...)):
    content = file.file.read()
    return data_import.import_style_config(content)


@router.post("/api/unproduced/upload")
def upload_unproduced(file: UploadFile = File(...)):
    content = file.file.read()
    return data_import.import_unproduced(content)


@router.post("/api/sales/upload")
def upload_sales(file: UploadFile = File(...)):
    content = file.file.read()
    return data_import.import_sales_raw(content)


@router.post("/api/inventory/upload")
def upload_inventory(file: UploadFile = File(...)):
    content = file.file.read()
    return data_import.import_inventory_raw(content)


@router.get("/api/lingxing/status")
def lingxing_status():
    conn = db.get_conn()
    cfg = lingxing._get_config(conn)
    conn.close()
    return {
        "configured": bool(cfg.get("lingxing_app_id") and cfg.get("lingxing_app_secret")),
        "app_id_set": bool(cfg.get("lingxing_app_id")),
        "app_secret_set": bool(cfg.get("lingxing_app_secret")),
        "host": cfg.get("lingxing_host") or lingxing.HOST,
        "sids": {k: v for k, v in cfg.items() if k.startswith("lingxing_sids_")},
    }


@router.post("/api/lingxing/config")
def set_lingxing_config(payload: dict):
    conn = db.get_conn()
    cur = conn.cursor()
    for key, value in payload.items():
        cur.execute(
            "INSERT INTO system_config (key, value, updated_at) VALUES (?, ?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
            (key, str(value), db.now()),
        )
    conn.commit()
    conn.close()
    return {"status": "ok"}


@router.post("/api/lingxing/sync")
def sync_lingxing(baseline_date: str = Form(...)):
    conn = db.get_conn()
    res = lingxing.sync_from_lingxing(conn, baseline_date)
    conn.close()
    return res


@router.post("/api/reports")
def create_report(name: str = Form(...), baseline_date: str = Form(...)):
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO reports (name, baseline_date, status, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
        (name, baseline_date, "pending", db.now(), db.now()),
    )
    report_id = cur.lastrowid
    conn.commit()
    conn.close()
    return {"status": "ok", "report_id": report_id}


@router.post("/api/reports/{report_id}/generate")
def generate_report(report_id: int, baseline_date: Optional[str] = Form(None)):
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute("SELECT baseline_date FROM reports WHERE id=?", (report_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="报告不存在")
    bd = baseline_date or row[0]
    return calc.generate_report(report_id, bd)


@router.get("/api/reports")
def list_reports():
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, name, baseline_date, status, created_at FROM reports ORDER BY created_at DESC"
    )
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    return {"items": rows}


@router.get("/api/reports/{report_id}/summary")
def report_summary(report_id: int):
    conn = db.get_conn()
    rows = calc.get_report_summary(conn, report_id)
    conn.close()
    return {"items": rows}


@router.get("/api/reports/{report_id}/detail")
def report_detail(report_id: int, platform: str, account: Optional[str] = None,
                  style_code: Optional[str] = None):
    conn = db.get_conn()
    if style_code:
        rows = calc.get_report_detail_by_style(conn, report_id, style_code, platform, account)
    else:
        rows = calc.get_report_detail(conn, report_id, platform, account)
    conn.close()
    cols = calc.detail_columns(platform, account)
    return {
        "columns": [{"label": lb, "field": f} for lb, f in cols],
        "items": rows,
    }


@router.get("/api/reports/{report_id}/structure")
def report_structure(report_id: int):
    """报表结构：款号列表 + 每个款号的批次列、汇总表头、平台分组（前端据此动态建表）"""
    conn = db.get_conn()
    styles = calc.get_report_styles(conn, report_id)
    out = []
    for sc in styles:
        batches = calc.get_style_batches(conn, sc)
        out.append({
            "style_code": sc,
            "batch_names": batches,
            "summary_columns": calc.summary_columns(batches),
            "detail_groups": calc.get_detail_groups(conn, report_id, sc),
        })
    conn.close()
    return {"styles": out}


@router.get("/api/reports/{report_id}/summary_table")
def report_summary_table(report_id: int, style_code: Optional[str] = None):
    """汇总表（按款号，含动态批次列），返回表头 + 二维数据，前端直接渲染"""
    conn = db.get_conn()
    styles = [style_code] if style_code else calc.get_report_styles(conn, report_id)
    out = []
    for sc in styles:
        batches = calc.get_style_batches(conn, sc)
        rows = calc.get_report_summary_by_style(conn, report_id, sc)
        out.append({
            "style_code": sc,
            "batch_names": batches,
            "columns": calc.summary_columns(batches),
            "rows": [calc.summary_row_values(r, batches) for r in rows],
        })
    conn.close()
    return {"styles": out}


@router.get("/api/config")
def get_config():
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute("SELECT key, value FROM system_config")
    cfg = dict(cur.fetchall())
    conn.close()
    return {
        "sellable_days_divisor": cfg.get("sellable_days_divisor", "denoised"),
        "sellable_days_rounding": cfg.get("sellable_days_rounding", "ceil"),
        "weights": calc.WEIGHTS,
    }


@router.post("/api/config")
def set_config(payload: dict):
    allowed = {"sellable_days_divisor", "sellable_days_rounding"}
    conn = db.get_conn()
    cur = conn.cursor()
    for k, v in payload.items():
        if k in allowed:
            cur.execute(
                "INSERT INTO system_config (key, value, updated_at) VALUES (?, ?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
                (k, str(v), db.now()),
            )
    conn.commit()
    conn.close()
    return {"status": "ok"}


@router.get("/api/reports/{report_id}/kpis")
def report_kpis(report_id: int):
    conn = db.get_conn()
    kpis = calc.get_kpis(conn, report_id)
    conn.close()
    return kpis


def _style_header(ws):
    """表头样式：加粗 + 灰底 + 冻结首行"""
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor="F2F2F2")
    for c in ws[1]:
        c.font = Font(bold=True, size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32
    for col in ws.columns:
        letter = col[0].column_letter
        width = max((len(str(c.value)) for c in col[:60] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 9), 26)


@router.post("/api/reports/{report_id}/export")
@router.get("/api/reports/{report_id}/export")
def export_report(report_id: int):
    """导出 Excel：结构对齐《各款号备货.xlsx》——每个款号 1 张汇总表 + N 张平台明细表"""
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute("SELECT name, baseline_date FROM reports WHERE id=?", (report_id,))
    meta = cur.fetchone()
    if not meta:
        conn.close()
        raise HTTPException(status_code=404, detail="报告不存在")
    report_name, baseline = meta

    styles = calc.get_report_styles(conn, report_id)
    if not styles:
        conn.close()
        raise HTTPException(status_code=404, detail="报告无数据，请先生成报表")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used = set()

    def _safe_title(t: str) -> str:
        t = "".join(ch for ch in t if ch not in "[]:*?/\\")[:31]
        base, i = t, 1
        while t in used:
            suffix = f"_{i}"
            t = base[: 31 - len(suffix)] + suffix
            i += 1
        used.add(t)
        return t

    for sc in styles:
        batches = calc.get_style_batches(conn, sc)
        # 汇总表
        ws = wb.create_sheet(title=_safe_title(f"{sc}备货汇总"))
        ws.append(calc.summary_columns(batches))
        for r in calc.get_report_summary_by_style(conn, report_id, sc):
            ws.append(calc.summary_row_values(r, batches))
        _style_header(ws)

        # 平台明细表
        for idx, g in enumerate(calc.get_detail_groups(conn, report_id, sc), start=1):
            cols = calc.detail_columns(g["platform"], g["account"])
            ws2 = wb.create_sheet(title=_safe_title(f"{g['label']}-输出表{idx}"))
            ws2.append([lb for lb, _ in cols])
            for r in calc.get_report_detail_by_style(
                conn, report_id, sc, g["platform"], g["account"]
            ):
                ws2.append([r.get(f) for _, f in cols])
            _style_header(ws2)

    # 附：原始数据表（便于核对）
    ws3 = wb.create_sheet(title=_safe_title("未生产数目-线下统计"))
    ws3.append(["款号", "sku", "批次", "未生产数目"])
    cur.execute(
        "SELECT style_code, sku, batch_name, quantity FROM unproduced_batches ORDER BY style_code, sku, batch_name"
    )
    for r in cur.fetchall():
        ws3.append(list(r))
    _style_header(ws3)

    conn.close()

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = quote(f"{report_name or '各款号备货'}_{baseline}.xlsx")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
