import io
import json
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import openpyxl
from fastapi import UploadFile

from . import database as db


def parse_excel(file_bytes: bytes, header_row: int = 1) -> List[List]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _find_sheet(wb, keywords: List[str]):
    """按关键词匹配 sheet 名"""
    for name in wb.sheetnames:
        for kw in keywords:
            if kw in name:
                return wb[name]
    return None


def import_style_config(file_bytes: bytes) -> Dict:
    """上传款号账号配置
    表头：款号, 生命周期, 亚马逊账号1, 备货周期, 配额, 亚马逊账号2, 备货周期, 配额, ...
    """
    rows = parse_excel(file_bytes)
    if not rows:
        return {"status": "error", "message": "空文件"}
    header = [str(h).strip() if h else "" for h in rows[0]]

    conn = db.get_conn()
    cur = conn.cursor()
    imported = 0

    for r in rows[1:]:
        if not r or not r[0]:
            continue
        style_code = str(r[0]).strip()
        lifecycle = str(r[1]).strip() if len(r) > 1 and r[1] else ""

        cur.execute(
            "INSERT INTO styles (style_code, lifecycle, created_at, updated_at) VALUES (?, ?, ?, ?) "
            "ON CONFLICT(style_code) DO UPDATE SET lifecycle=excluded.lifecycle, updated_at=excluded.updated_at",
            (style_code, lifecycle, db.now(), db.now()),
        )

        # 删除旧账号配置，重新写入
        cur.execute("DELETE FROM style_accounts WHERE style_code=?", (style_code,))

        # 按 (账号, 备货周期, 配额) 三元组解析
        idx = 2
        sort_order = 0
        while idx + 1 < len(r):
            account = str(r[idx]).strip() if r[idx] else ""
            if not account:
                idx += 3
                continue
            days = int(r[idx + 1]) if len(r) > idx + 1 and r[idx + 1] is not None else 134
            quota = float(r[idx + 2]) if len(r) > idx + 2 and r[idx + 2] is not None else 0
            cur.execute(
                "INSERT INTO style_accounts (style_code, platform, account_name, account_prefix, stocking_days, quota, sort_order, created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    style_code,
                    "amazon",
                    account,
                    account[:2].upper(),
                    days,
                    quota,
                    sort_order,
                    db.now(),
                    db.now(),
                ),
            )
            sort_order += 1
            idx += 3

        # 默认补充沃尔玛 / 其他 / temu（无配额，默认备货周期 134）
        for platform in ["walmart", "other", "temu"]:
            cur.execute(
                "INSERT INTO style_accounts (style_code, platform, account_name, account_prefix, stocking_days, quota, sort_order, created_at, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) "
                "ON CONFLICT(style_code, platform, account_name) DO NOTHING",
                (style_code, platform, platform, platform[:2].upper(), 134, 0, sort_order, db.now(), db.now()),
            )
            sort_order += 1

        imported += 1

    conn.commit()
    conn.close()
    return {"status": "ok", "imported_styles": imported}


def import_unproduced(file_bytes: bytes) -> Dict:
    """上传未生产数目
    每个 sheet 名为款号；列：sku, 批次1, 批次2, ..., 总未生产数目
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    conn = db.get_conn()
    cur = conn.cursor()
    imported_sheets = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        style_code = str(sheet_name).strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h else "" for h in rows[0]]
        # 删除该款式旧数据
        cur.execute("DELETE FROM unproduced_batches WHERE style_code=?", (style_code,))

        batch_cols = header[1:-1]  # 假设最后一列是总未生产数目
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            sku = str(r[0]).strip()
            for i, col_name in enumerate(batch_cols):
                val = r[i + 1] if len(r) > i + 1 else 0
                qty = float(val) if val is not None else 0
                cur.execute(
                    "INSERT INTO unproduced_batches (style_code, sku, batch_name, quantity, uploaded_at) VALUES (?, ?, ?, ?, ?)",
                    (style_code, sku, col_name, qty, db.now()),
                )
        imported_sheets += 1

    conn.commit()
    conn.close()
    wb.close()
    return {"status": "ok", "imported_sheets": imported_sheets}


def _classify_platform_label(label: str):
    label = str(label)
    if "ZO" in label:
        return "amazon", "ZO"
    if "PS" in label:
        return "amazon", "PS"
    if "HW" in label:
        return "amazon", "HW"
    if "SZ" in label:
        return "amazon", "SZ"
    if "其他" in label:
        return "other", None
    if "沃尔玛" in label:
        return "walmart", None
    if "temu" in label or "Temu" in label or "TEMU" in label:
        return "temu", None
    if "中国仓" in label:
        return "other", None
    return "other", None


def _parse_date(date_raw) -> str:
    if isinstance(date_raw, datetime):
        return date_raw.strftime("%Y%m%d")
    if date_raw is None or date_raw == "":
        return ""
    s = str(date_raw).strip().replace("-", "").replace("/", "")
    if s.isdigit() and len(s) == 8:
        return s
    # 尝试解析 ISO 日期 20260620 -> 已经满足
    try:
        dt = datetime.strptime(s, "%Y%m%d")
        return dt.strftime("%Y%m%d")
    except Exception:
        pass
    return ""


def import_sales_raw(file_bytes: bytes) -> Dict:
    """上传销售数据-领星获取（兜底）
    表头行格式见样例：第一行平台分组，第二行子列名
    平台顺序：ZO / PS / HW / 其他 / 沃尔玛 / temu
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _find_sheet(wb, ["销售数据", "销售", "sales"]) or wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return {"status": "error", "message": "销售表数据不足"}

    platform_row = [str(h).strip() if h else "" for h in rows[0]]
    header_row = [str(h).strip() if h else "" for h in rows[1]]
    segments = []
    for i, h in enumerate(header_row):
        if h in ["SKU", "sku"]:
            platform_label = ""
            for j in range(i, -1, -1):
                if platform_row[j] and platform_row[j] not in ["", "账号"]:
                    platform_label = platform_row[j]
                    break
            platform, account = _classify_platform_label(platform_label)
            # temu 区段只有 sku + 近30天销量两列
            is_temu = platform == "temu"
            qty_col = i + 1
            date_col = None if is_temu else i - 1
            segments.append({"sku_col": i, "qty_col": qty_col, "date_col": date_col, "platform": platform, "account": account, "is_temu": is_temu})

    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM sales_raw")
    imported = 0
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")

    for r in rows[2:]:
        for seg in segments:
            if len(r) <= seg["sku_col"]:
                continue
            sku = r[seg["sku_col"]]
            qty = r[seg["qty_col"]] if len(r) > seg["qty_col"] else None
            if not sku or qty is None or qty == "":
                continue
            sku = str(sku).strip()
            qty = float(qty)

            if seg["is_temu"]:
                date_str = yesterday
            else:
                date_raw = r[seg["date_col"]] if seg["date_col"] is not None and len(r) > seg["date_col"] else None
                date_str = _parse_date(date_raw)
                if not date_str:
                    continue
            cur.execute(
                "INSERT INTO sales_raw (date, sku, platform, account, quantity, fetched_at) VALUES (?, ?, ?, ?, ?, ?)",
                (date_str, sku, seg["platform"], seg["account"], qty, db.now()),
            )
            imported += 1

    conn.commit()
    conn.close()
    return {"status": "ok", "imported_rows": imported}


def import_inventory_raw(file_bytes: bytes) -> Dict:
    """上传库存数据-领星获取（兜底）
    表头行格式见样例：第一行平台分组，第二行子列名
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _find_sheet(wb, ["库存数据", "库存", "inventory"]) or wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return {"status": "error", "message": "库存表数据不足"}

    platform_row = [str(h).strip() if h else "" for h in rows[0]]
    header_row = [str(h).strip() if h else "" for h in rows[1]]
    segments = []
    for i, h in enumerate(header_row):
        if h in ["SKU", "sku"]:
            platform_label = ""
            for j in range(i, -1, -1):
                if platform_row[j] and platform_row[j] not in ["", "账号"]:
                    platform_label = platform_row[j]
                    break
            platform, account = _classify_platform_label(platform_label)
            warehouse = header_row[i + 1] if len(header_row) > i + 1 else platform_label
            segments.append({"sku_col": i, "qty_col": i + 1, "platform": platform, "account": account, "warehouse": warehouse})

    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM inventory_raw")
    imported = 0

    for r in rows[2:]:
        for seg in segments:
            if len(r) <= seg["sku_col"]:
                continue
            sku = r[seg["sku_col"]]
            qty = r[seg["qty_col"]] if len(r) > seg["qty_col"] else None
            if not sku or qty is None or qty == "":
                continue
            sku = str(sku).strip()
            qty = float(qty)
            cur.execute(
                "INSERT INTO inventory_raw (sku, platform, account, warehouse, quantity, fetched_at) VALUES (?, ?, ?, ?, ?, ?)",
                (sku, seg["platform"], seg["account"], seg["warehouse"], qty, db.now()),
            )
            imported += 1

    conn.commit()
    conn.close()
    return {"status": "ok", "imported_rows": imported}
